import math
import os, sys
import shutil
from openpyxl import load_workbook
divid=int(input("input N wanted: "))
log=open("log.txt", 'a+')
from datetime import datetime
now = datetime.now()
current_time = now.strftime("%H:%M:%S")
log.write("#######################################\n")
log.write("Start processing: "+current_time+"\n")
log.write("the value of N: "+str(divid)+"\n")
log.close()
number_array=[]
for n1 in range(0,divid):
    for n2 in range(0,divid):
        for n3 in range(0,divid):
            for n4 in range(0,divid):
                number_array.append([n1,n2,n3,n4])
none_existed=0
try:
    os.mkdir('N='+str(divid))
except:
    none_existed=1
lists=[(1,1,1,1),(1,1,1,2),(1,1,1,3),(1,1,1,4),(1,1,1,5),(1,1,2,2),(1,1,2,3),(1,1,2,4),(1,1,2,6),(1,1,3,3),(1,2,2,2),(1,2,2,3),(1,2,2,4),(1,2,2,6),(1,2,4,4),(1,2,4,6)]
done_number=0
done_names=[]
if none_existed==1:
    for file in os.listdir('N='+str(divid)+'/txt'):
        done_number+=1
        done_names.append(file)
    print("have already finished: "+str(done_names))
    for i in range(done_number):
        lists.pop(0)
else:
    print("have no database, creating")
for k in range(len(lists)):
    try:
        os.mkdir('N='+str(divid)+"/txt")
    except:
        noti=0
    A='N='+str(divid)+'/txt/'+str(lists[k])+'/'
    try:
        os.mkdir(A)
    except:
        a='skip'
    
    print(str(k)+"...")
    for i in range((divid)**4):
        h=number_array[i]
        hname=os.path.join(A,('h'+'='+str(h))+".txt")
        fas = open(hname,'w')
        s = [] # s[i] represents the number of solutions for n = i
        
        a = lists[k]
        from ast import literal_eval
        fin = open('metadata/meta'+str(a)+'.txt','r')
        lines = fin.readlines()
        ans = [] # ans[i] represents the number of solutions for n = i

        N = divid #mod number

        def judge(x1,x2,x3,x4):
            global N, h
            if x1 % N == h[0] and x2 % N == h[1] and x3 % N == h[2] and x4 % N == h[3]:
                return True
            else:
                return False

        n = 0
        result = 0
        for line in lines:
            if line == '*\n':
                n += 1
                ans.append(result)
                result = 0
                continue
            line = literal_eval(line[:-1])
            x1 = int(line[0])
            x2 = int(line[1])
            x3 = int(line[2])
            x4 = int(line[3])
            if x1 == 0:
                if x2 == 0:
                    if x3 == 0:
                        if x4 == 0:
                            if judge(x1, x2, x3, x4):
                                result += 1
                        else:  # x4 != 0
                            if judge(x1, x2, x3, x4):
                                result += 1
                            if judge(x1, x2, x3, -x4):
                                result += 1
                    else:  # x3 != 0
                        if x4 == 0:
                            if judge(x1, x2, x3, x4):
                                result += 1
                            if judge(x1, x2, -x3, x4):
                                result += 1
                        else:  # x4 != 0
                            if judge(x1, x2, x3, x4):
                                result += 1
                            if judge(x1, x2, x3, -x4):
                                result += 1
                            if judge(x1, x2, -x3, x4):
                                result += 1
                            if judge(x1, x2, -x3, -x4):
                                result += 1
                else:  # x2 != 0
                    if x3 == 0:
                        if x4 == 0:
                            if judge(x1, x2, x3, x4):
                                result += 1
                            if judge(x1, -x2, x3, x4):
                                result += 1
                        else:  # x4 != 0
                            if judge(x1, x2, x3, x4):
                                result += 1
                            if judge(x1, x2, x3, -x4):
                                result += 1
                            if judge(x1, -x2, x3, x4):
                                result += 1
                            if judge(x1, -x2, x3, -x4):
                                result += 1
                    else:  # x3 != 0
                        if x4 == 0:
                            if judge(x1, x2, x3, x4):
                                result += 1
                            if judge(x1, x2, -x3, x4):
                                result += 1
                            if judge(x1, -x2, x3, x4):
                                result += 1
                            if judge(x1, -x2, -x3, x4):
                                result += 1
                        else:  # x4 != 0
                            if judge(x1, x2, x3, x4):
                                result += 1
                            if judge(x1, x2, x3, -x4):
                                result += 1
                            if judge(x1, x2, -x3, x4):
                                result += 1
                            if judge(x1, x2, -x3, -x4):
                                result += 1
                            if judge(x1, -x2, x3, x4):
                                result += 1
                            if judge(x1, -x2, x3, -x4):
                                result += 1
                            if judge(x1, -x2, -x3, x4):
                                result += 1
                            if judge(x1, -x2, -x3, -x4):
                                result += 1
            else:  # x1 != 0
                if x2 == 0:
                    if x3 == 0:
                        if x4 == 0:
                            if judge(x1, x2, x3, x4):
                                result += 1
                            if judge(-x1, x2, x3, x4):
                                result += 1
                        else:  # x4 != 0
                            if judge(x1, x2, x3, x4):
                                result += 1
                            if judge(x1, x2, x3, -x4):
                                result += 1
                            if judge(-x1, x2, x3, x4):
                                result += 1
                            if judge(-x1, x2, x3, -x4):
                                result += 1
                    else:  # x3 != 0
                        if x4 == 0:
                            if judge(x1, x2, x3, x4):
                                result += 1
                            if judge(x1, x2, -x3, x4):
                                result += 1
                            if judge(-x1, x2, x3, x4):
                                result += 1
                            if judge(-x1, x2, -x3, x4):
                                result += 1
                        else:  # x4 != 0
                            if judge(x1, x2, x3, x4):
                                result += 1
                            if judge(x1, x2, x3, -x4):
                                result += 1
                            if judge(x1, x2, -x3, x4):
                                result += 1
                            if judge(x1, x2, -x3, -x4):
                                result += 1
                            if judge(-x1, x2, x3, x4):
                                result += 1
                            if judge(-x1, x2, x3, -x4):
                                result += 1
                            if judge(-x1, x2, -x3, x4):
                                result += 1
                            if judge(-x1, x2, -x3, -x4):
                                result += 1
                else:  # x2 != 0
                    if x3 == 0:
                        if x4 == 0:
                            if judge(x1, x2, x3, x4):
                                result += 1
                            if judge(x1, -x2, x3, x4):
                                result += 1
                            if judge(-x1, x2, x3, x4):
                                result += 1
                            if judge(-x1, -x2, x3, x4):
                                result += 1
                        else:  # x4 != 0
                            if judge(x1, x2, x3, x4):
                                result += 1
                            if judge(x1, x2, x3, -x4):
                                result += 1
                            if judge(x1, -x2, x3, x4):
                                result += 1
                            if judge(x1, -x2, x3, -x4):
                                result += 1
                            if judge(-x1, x2, x3, x4):
                                result += 1
                            if judge(-x1, x2, x3, -x4):
                                result += 1
                            if judge(-x1, -x2, x3, x4):
                                result += 1
                            if judge(-x1, -x2, x3, -x4):
                                result += 1
                    else:  # x3 != 0
                        if x4 == 0:
                            if judge(x1, x2, x3, x4):
                                result += 1
                            if judge(x1, x2, -x3, x4):
                                result += 1
                            if judge(x1, -x2, x3, x4):
                                result += 1
                            if judge(x1, -x2, -x3, x4):
                                result += 1
                            if judge(-x1, x2, x3, x4):
                                result += 1
                            if judge(-x1, x2, -x3, x4):
                                result += 1
                            if judge(-x1, -x2, x3, x4):
                                result += 1
                            if judge(-x1, -x2, -x3, x4):
                                result += 1
                        else:  # x4 != 0
                            if judge(x1, x2, x3, x4):
                                result += 1
                            if judge(x1, x2, x3, -x4):
                                result += 1
                            if judge(x1, x2, -x3, x4):
                                result += 1
                            if judge(x1, x2, -x3, -x4):
                                result += 1
                            if judge(x1, -x2, x3, x4):
                                result += 1
                            if judge(x1, -x2, x3, -x4):
                                result += 1
                            if judge(x1, -x2, -x3, x4):
                                result += 1
                            if judge(x1, -x2, -x3, -x4):
                                result += 1
                            if judge(-x1, x2, x3, x4):
                                result += 1
                            if judge(-x1, x2, x3, -x4):
                                result += 1
                            if judge(-x1, x2, -x3, x4):
                                result += 1
                            if judge(-x1, x2, -x3, -x4):
                                result += 1
                            if judge(-x1, -x2, x3, x4):
                                result += 1
                            if judge(-x1, -x2, x3, -x4):
                                result += 1
                            if judge(-x1, -x2, -x3, x4):
                                result += 1
                            if judge(-x1, -x2, -x3, -x4):
                                result += 1
        for i in range(0, n):
            fas.writelines(str(ans[i]))
            fas.write('\n')
        fas.close()
log=open("log.txt", 'a+')
log.write("Database created or imported\n")
log.write("==============================================\n")
log.write("start excel compiling\n")
log.close()
print("==============================================")
print("start excel compiling")
import os

LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
def excel_style(col):
    """ Convert given row and column number to an Excel-style cell name. """
    result = []
    while col:
        col, rem = divmod(col-1, 26)
        result[:0] = LETTERS[rem]
    return ''.join(result)
excel_names=[]
for i in range(2*(divid**4+10)):
    excel_names.append(excel_style(i+1))
listr=[(1,1,1,1),(1,1,1,2),(1,1,1,3),(1,1,1,4),(1,1,1,5),(1,1,2,2),(1,1,2,3),(1,1,2,4),(1,1,2,6),(1,1,3,3),(1,2,2,2),(1,2,2,3),(1,2,2,4),(1,2,2,6),(1,2,4,4),(1,2,4,6)]


from openpyxl import Workbook
import openpyxl


for k in range(16):
    names=[]
    datas=[]
    data=[]
    number=0
    for file in os.listdir('N='+str(divid)+'/txt/'+str(listr[k])):
        number+=1
        names.append(file[:-4])
    print(number)
    print(listr[k])
    try:
        book = load_workbook('N='+str(divid)+'.xlsx')
    except:
        wb = Workbook()
        wb.save('N='+str(divid)+'.xlsx')
        book = load_workbook(filename='N='+str(divid)+'.xlsx')
    sheet = book.create_sheet(str(listr[k]))

    sheet['A1'] = "n"
    for i in range(2,103):
        write1="A"+str(i+1)
        sheet[write1]=str(i-2)
    sheet['B1'] = "a="+str(listr[k])
    filename=str(listr[k])+".txt"
    file=open(filename, "r+")
    with open(filename, "r") as fps:
        datas = fps.readlines()
    for i in range(2,103):
        write6="B"+str(i)
        sheet[write6]=datas[i-1]
    file.close()

    for i in range(divid**4):
        write3=str(excel_names[2*(i+1)])+str(1)
        write4=str(names[i])
        sheet[write3]=write4
        txtname='N='+str(divid)+"/txt/"+str(listr[k])+"/"+str(names[i])+".txt"
        file1 = open(txtname,"r+")
        with open(txtname, "r") as fp:
            data = fp.readlines()
        for j in range(2,103):
            write2=str(excel_names[2*(i+1)])+str(j)
            sheet[write2]=data[j-2][:-1]
            write7=str(excel_names[2*(i+1)+1])+str(j)
            if int(data[j-2][:-1])!=0:
                sheet[write7]=str(int(datas[j-1])/int(data[j-2][:-1]))
            else:
                sheet[write7]="null"

    
    log=open("log.txt", 'a+')
    log.write("Data for sheet"+str(k+1)+" is written and back-up created\n")
    log.write("Name for sheet is: "+str(listr[k]))
    log.write("\n-------------------------------------------------\n")
    log.close()
    file1.close()
    book.save('N='+str(divid)+'.xlsx')
    original='N='+str(divid)+'.xlsx'
    target='datastorage_'+original
    shutil.copyfile(original, target)
shutil.move(('N='+str(divid)+'.xlsx'),('N='+str(divid)))
