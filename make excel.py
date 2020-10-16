import os

excel_names=['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ', 'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ', 'CR', 'CS', 'CT', 'CU', 'CV', 'CW', 'CX', 'CY', 'CZ', 'DA', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DH', 'DI', 'DJ', 'DK', 'DL', 'DM', 'DN', 'DO', 'DP', 'DQ', 'DR', 'DS', 'DT', 'DU', 'DV', 'DW', 'DX', 'DY', 'DZ', 'EA', 'EB', 'EC', 'ED', 'EE', 'EF', 'EG', 'EH', 'EI', 'EJ', 'EK', 'EL', 'EM', 'EN', 'EO', 'EP', 'EQ', 'ER', 'ES', 'ET', 'EU', 'EV', 'EW', 'EX', 'EY', 'EZ', 'FA', 'FB', 'FC', 'FD', 'FE', 'FF', 'FG', 'FH', 'FI', 'FJ', 'FK', 'FL', 'FM', 'FN', 'FO', 'FP', 'FQ', 'FR', 'FS', 'FT', 'FU', 'FV', 'FW', 'FX', 'FY', 'FZ', 'GA', 'GB', 'GC', 'GD', 'GE', 'GF', 'GG', 'GH', 'GI', 'GJ', 'GK', 'GL', 'GM', 'GN', 'GO', 'GP', 'GQ', 'GR', 'GS', 'GT', 'GU', 'GV', 'GW', 'GX', 'GY', 'GZ']

listr=[(1,1,1,1),(1,1,1,2),(1,1,1,3),(1,1,1,4),(1,1,1,5),(1,1,2,2),(1,1,2,3),(1,1,2,4),(1,1,2,6),(1,1,3,3),(1,2,2,2),(1,2,2,3),(1,2,2,4),(1,2,2,6),(1,2,4,4),(1,2,4,6)]


from openpyxl import Workbook
import openpyxl
divid=2
for k in range(16):
    print(listr[k])
    names=[]
    datas=[]
    data=[]
    number=0
    for file in os.listdir('txt/'+str(listr[k])):
        number+=1
        names.append(file[:-4])
    print(names)
    print(number)
    try:
        book = openpyxl.load_workbook('N='+str(divid)+'.xlsx')
    except:
        wb = Workbook()
        wb.save('N='+str(divid)+'.xlsx')
        book = openpyxl.load_workbook('N='+str(divid)+'.xlsx')
    sheet = book.create_sheet(str(listr[k]))

    sheet['A1'] = "n"
    for i in range(2,1003):
        write1="A"+str(i)
        sheet[write1]=str(i-1)
    sheet['B1'] = "a="+str(listr[k])
    filename=str(listr[k])+".txt"
    file=open(filename, "r+")
    with open(filename, "r") as fps:
        datas = fps.readlines()
    for i in range(2,1003):
        write6="B"+str(i)
        sheet[write6]=datas[i-1]
    file.close()

    for i in range(81):
        write3=str(excel_names[2*(i+1)])+str(1)
        write4=str(names[i])
        sheet[write3]=write4
        txtname="txt/"+str(listr[k])+"/"+str(names[i])+".txt"
        file1 = open(txtname,"r+")
        with open(txtname, "r") as fp:
            data = fp.readlines()
        for j in range(2,1003):
            write2=str(excel_names[2*(i+1)])+str(j)
            sheet[write2]=data[j-2][:-1]
            write7=str(excel_names[2*(i+1)+1])+str(j)
            if int(data[j-2][:-1])!=0:
                sheet[write7]=str(int(datas[j-1])/int(data[j-2][:-1]))
            else:
                sheet[write7]="null"
    book.save('N='+str(divid)+'.xlsx')
    file1.close()
